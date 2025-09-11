from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
import os
# Carga de variables de entorno desde .envy listo
from dotenv import load_dotenv
load_dotenv()
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from flask_migrate import Migrate
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


#C:\Users\Guillermo\AppData\Local\Programs\Python\Python313\python.exe "$(FULL_CURRENT_PATH)"


def convertir_hora_a_decimal(hora_str):
    try:
        return float(int(hora_str.strip()))
    except ValueError:
        return 0.0



app = Flask(__name__)
app.secret_key = 'clave_secreta_para_sesiones'

app.config['TEMPLATES_AUTO_RELOAD'] = True
app.jinja_env.cache = {}

# Configuración para PostgreSQL
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get("DATABASE_URL")
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# Inicializar Flask-Migrate
migrate = Migrate(app, db)
migrate.init_app(app, db)

# ─── Modelos ─────────────────────────────────────
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(50), nullable=False)

    registros = db.relationship('Registro', backref='user', lazy=True)

class Registro(db.Model):
    __tablename__ = 'registros'

    id              = db.Column(db.Integer, primary_key=True)
    user_id         = db.Column(db.Integer, db.ForeignKey('users.id'))
    fecha           = db.Column(db.String(50))
    entrada         = db.Column(db.String(50))
    salida          = db.Column(db.String(50))
    almuerzo        = db.Column(db.Float)
    viaje_ida       = db.Column(db.Float, default=0)
    viaje_vuelta    = db.Column(db.Float, default=0)
    km_ida          = db.Column(db.Float, default=0)
    km_vuelta       = db.Column(db.Float, default=0)
    horas           = db.Column(db.Float)
    tarea           = db.Column(db.Text)
    cliente         = db.Column(db.Text)
    comentarios     = db.Column(db.Text)
    contrato = db.Column(db.String(100))
    service_order = db.Column(db.String(100))
    centro_costo = db.Column(db.String(100))  # ← Simple texto, no ID
    tipo_servicio = db.Column(db.String(100))  # ← Texto
    linea = db.Column(db.String(100))  # ← Texto

    # (opcional) si querés acceder al usuario desde el registro:
    # user = db.relationship('User', backref='registros')


# ─── Inicialización de la base de datos ─────────
with app.app_context():
    db.create_all()
    if not User.query.filter(db.func.lower(User.username) == 'guillermo gutierrez').first():
        superadmin = User(username='guillermo gutierrez', password='0000', role='superadmin')
        db.session.add(superadmin)
        db.session.commit()

# ─── Rutas ──────────────────────────────────────
@app.route('/', methods=['GET', 'POST'])
def inicio():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username'].strip().lower()
        password = request.form['password']

        user = User.query.filter(
            db.func.lower(User.username) == username,
            User.password == password
        ).first()

        if user:
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            return redirect(url_for('dashboard'))
        else:
            flash('Usuario o contraseña incorrectos', category='danger')
    return render_template('login.html')
    

@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    tipo_cliente = request.args.get('tipo_cliente', 'Todos')
    
    # Ejemplo de listas de opciones (reemplazar por consulta a DB luego)
    
    clientes = [
        '1561046 - Barraca Deambrosi SA (Carton)',
        '1561046 - Barraca Deambrosi SA (Proceso)',
        '1561458 - Belficor Sociedad Anónima (Proceso)',
        '1561459 - Belficor - Paraje La Boyada',
        '1560018 - Cerealin San Jose (Carton)',
        '1560018 - Cerealin San Jose (Proceso)',
        '1569004 - Cerealin S.A.',
        '1561233 - Companía Salus SA',
        '1561245 - Conaprole Planta I (Proceso)',
        '1561250 - Conaprole Planta CIM (Carton)',
        '1561250 - Conaprole Planta CIM (Proceso)',
        '1561010 - Conaprole Planta VII (Proceso)',
        '1560606 - Conaprole Planta VIII (Carton)',
        '1560606 - Conaprole Planta VIII (Proceso)',
        '1561251 - Cooperativa Agraria (CALCAR)',
        '1560621 - Ecolat Uruguay S.A. (Proceso)',
        '1561335 - Gibur S.A. (Carton)',
        '1561062 - Jugos del Uruguay S.A. (Carton)',
        '1561062 - Jugos del Uruguay S.A. (Proceso)',
        '1561132 - Montevideo Refrescos S.A. (Proceso)',
        '1561540 - Nolir S.A.',
        '1560020 - Recalco S.A. (ex Suadil)',
        '1570011 - Tetrapak Argentina',
        'Otros Clientes del Mercado'
    ]
    if tipo_cliente == 'Cartón':
        clientes = [c for c in clientes if cliente_prefijo.get(c, '').startswith('UYC')]
    elif tipo_cliente == 'Proceso':
        clientes = [c for c in clientes if cliente_prefijo.get(c, '').startswith('UYP')]


    service_orders = ['SM02', 'SM03','N/A']
    centros_costo = [
        {'id': 1, 'nombre': '1561046 - Barraca Deambrosi SA (Carton) C.Costo=1 ()'},
        {'id': 2, 'nombre': '1561251 - Cooperativa Agraria (CALCAR) C.Costo=2 ()'},
        {'id': 3, 'nombre': '1561335 - Gibur S.A. (Carton) C.Costo=3 ()'},
        {'id': 4, 'nombre': '1561540 - Nolir S.A. C.Costo=4 ()'},
        {'id': 5, 'nombre': '1560020 - Recalco S.A. (ex Suadil) C.Costo=5 ()'},
        {'id': 6, 'nombre': '1561250 - Conaprole Planta CIM (Carton) C.Costo=6 ()'},
        {'id': 7, 'nombre': '1561245 - Conaprole Planta I (Proceso) C.Costo= ()'},
        {'id': 8, 'nombre': '1561250 - Conaprole Planta CIM (Proceso) C.Costo= ()'},
        {'id': 9, 'nombre': '1561010 - Conaprole Planta VII (Proceso) C.Costo= ()'},
        {'id': 10, 'nombre': '1560606 - Conaprole Planta VIII (Carton) C.Costo=7 ()'},
        {'id': 11, 'nombre': '1560606 - Conaprole Planta VIII (Proceso) C.Costo= ()'},
        {'id': 12, 'nombre': '1560018 -Cerealin San Jose (Carton)C.Costo=8 ()'},
        {'id': 13, 'nombre': '1560018 -Cerealin San Jose (Proceso) C.Costo=8 ()'},
        {'id': 14, 'nombre': '1561062 - Jugos del Uruguay S.A. (Carton) GMB revisar ()'},
        {'id': 15, 'nombre': '1561062 - Jugos del Uruguay S.A. (Proceso) GMB revisar ()'},
        {'id': 16, 'nombre': 'FUERA DE CONTRATO'},       
        {'id': 17, 'nombre': '1561458 - Belficor Sociedad Anónima (Proceso) C.Costo='},
        {'id': 18, 'nombre': '1561459 - Belficor - Paraje La Boyada C.Costo='},
        {'id': 19, 'nombre': '1561233 - Companía Salus SA C.Costo='},
        {'id': 20, 'nombre': 'Barraca Deambrosi SA (Proceso) C.Costo=1 ()'},
        {'id': 22, 'nombre': '1561132 - Montevideo Refrescos S.A. (Proceso) C.Costo= ()'},
        {'id': 23, 'nombre': '1569004 - Cerealin S.A. C.Costo= ()'},
        {'id': 24, 'nombre': '1560621 - Ecolat Uruguay S.A. (Proceso) C.Costo= ()'},
        {'id': 25, 'nombre': 'Otros Clientes del Mercado N/A'}
    ]
    contratos = [
        {'value': '73450003', 'label': 'Contrato'},
        {'value': '79010000', 'label': 'Administrativo - 79010000'},
        {'value': '79200030', 'label': 'Issue Resolution - Service - 79200030'},
        {'value': '79200020', 'label': 'Issue Resolution - Capital CARTON - 79200020'},
        {'value': '79200050', 'label': 'Issue Resolution - Capital PROCESSING - 79200050'},
        {'value': '580000', 'label': 'Capital Equipment carton - 580000'},
        {'value': '70350000', 'label': 'Capital Equipment processing - 70350000'},
        {'value': '73450000', 'label': 'Fuera de contrato - 73450000'},
        {'value': '560000', 'label': 'Upgrade projects - 560000'},
        {'value': '70158000', 'label': 'Mandatory kit - 70158000'},
        {'value': '480000', 'label': 'Rebilling - 480000'},
        {'value': '73453000', 'label': 'Training Facturable - 73453000'},
        {'value': '79010000', 'label': 'Training Interno - 79010000'},
        {'value': '70400010', 'label': 'MDR - 70400010'},
        {'value': '79010000', 'label': 'No disponible - 79010000'},
        {'value': '0', 'label': 'N/A'}
    ]

    tipos_servicio = [
        {'id': 1, 'nombre': 'TPMS'},
        {'id': 2, 'nombre': 'Correctivo'},
        {'id': 3, 'nombre': 'Soporte a la producción'},
        {'id': 4, 'nombre': 'Proyectos'},
        {'id': 5, 'nombre': 'RK'},
        {'id': 6, 'nombre': 'Vacaciones'},
        {'id': 7, 'nombre': 'Tareas Administrativas'},
        {'id': 8, 'nombre': 'Compensatorios'},
        {'id': 9, 'nombre': 'Training'},
        {'id': 10, 'nombre': 'Claims'}
    ]
    lineas = [
        {'id': 1,  'nombre': 'UYC-BARRACA   MVD-LIN01   Máquina-TBA/3                               N/S-11443/05537'},
        {'id': 2,  'nombre': 'UYC-BARRACA   MVD-LIN02   Máquina-TBA/8                               N/S-20201/82004'},
        {'id': 3,  'nombre': 'UYC-BARRACA   MVD-LIN03   Máquina-SIMPLY8                             N/S-21222/00018'},
        {'id': 4,  'nombre': 'UYC-BARRACA   MVD-LIN04   Máquina-TBA/19                              N/S-20562/83308'},
        {'id': 5,  'nombre': 'UYC-COAGRARIA CAR-LN 01   Máquina-TBA/8                               N/S-13037/10830'},
        {'id': 6,  'nombre': 'UYC-COAGRARIA CAR-LN 02   Máquina-TP C3/F                             N/S-15034/00004'},
        {'id': 7,  'nombre': 'UYC-NOLIR     MVD-LIN01   Máquina-TBA/19                              N/S-20591/83337'},
        {'id': 8,  'nombre': 'UYC-NOLIR     MVD-LIN02   Máquina-TBA/8                               N/S-15010/00889'},
        {'id': 9,  'nombre': 'UYC-CEREALIN  SJO-LIN01   Máquina-TBA/8                               N/S-13588/11417'},
        {'id': 10, 'nombre': 'UYC-CEREALIN  SJO-LIN04   Máquina-TP A3/CF                            N/S-21220/00466'},
        {'id': 11, 'nombre': 'UYC-CONAPROLE CIM-LIN02   Máquina-TBA/19                              N/S-20258/82571'},
        {'id': 12, 'nombre': 'UYC-CONAPROLE CIM-LIN03   Máquina-TT/3                                N/S-63202/20090'},
        {'id': 13, 'nombre': 'UYC-CONAPROLE P08-LIN01   Máquina-TBA/8                               N/S-20239/82382'},
        {'id': 14, 'nombre': 'UYC-CONAPROLE P08-LIN02   Máquina-TBA/8                               N/S-13879/11665'},
        {'id': 15, 'nombre': 'UYC-CONAPROLE P08-LIN03   Máquina-TBA/8                               N/S-13457/11304'},
        {'id': 16, 'nombre': 'UYC-CONAPROLE P08-LIN04   Máquina-TBA/8                               N/S-13486/11332'},
        {'id': 17, 'nombre': 'UYC-GIBUR     MVD-LIN01   Máquina-TBA/8                               N/S-17010/00018'},
        {'id': 18, 'nombre': 'UYC-GIBUR     MVD-LIN02   Máquina-TBA/19                               N/S-21200/00201'},
        {'id': 19, 'nombre': 'UYC-RECALCO   MVD-LIN01   Máquina-TBA/3                               N/S-20078/80780'},
        {'id': 20, 'nombre': 'UYC-RECALCO   MVD-LIN02   Máquina-TBA/8                               N/S-12967/10664'},
        {'id': 21, 'nombre': 'UYP-BELFICOR  MDV-CIP     Máquina-Tetra Alcip 2                       N/S-C6075349/01'},
        {'id': 22, 'nombre': 'UYP-FTE SALUS MIN-TREAT   Máquina-Tetra Therm Nacional                N/S-B04013-02'},
        {'id': 23, 'nombre': 'UYP-FTE SALUS MIN-TREAT   Máquina-Tetra Plex MS6 SR                   N/S-30105-08295'},
        {'id': 24, 'nombre': 'UYP-FTE SALUS MIN-TREAT   Máquina-Tetra Plex C8 in B6069919/01        N/S-30106-34366'},
        {'id': 25, 'nombre': 'UYP-FTE SALUS MIN-TREAT   Máquina-Tetra Therm Nacional                N/S-B6069919/01'},
        {'id': 26, 'nombre': 'UYP-FTE SALUS MIN-TREAT   Máquina-Tetra Alcip Nacional                N/S-B6069919/02'},
        {'id': 27, 'nombre': 'UYP-FTE SALUS MIN-TREAT   Máquina-Tanque Ultra Clean Nacional         N/S-B6123596/01'},
        {'id': 28, 'nombre': 'UYP-FTE SALUS MIN-TREAT   Máquina-Modulo CIP                          N/S-B6249678/01 '},
        {'id': 29, 'nombre': 'UYP-CONAPROLE CIM-IC 01   Máquina-Tetra Pak® Ingredient Doser 4000 A2 N/S-Z2014162 '},
        {'id': 30, 'nombre': 'UYP-CONAPROLE CIM-IC 01   Máquina-Choice Filler by Tetra Pak®         N/S-395-008-2019 '},
        {'id': 31, 'nombre': 'UYP-CONAPROLE CIM-MP 01   Máquina-Tetra Pak Aseptic Tank VD           N/S-T5844611780 '},
        {'id': 32, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-TETRA ALEX 30                       N/S-T5845523408 '},
        {'id': 33, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-TETRA ALEX 30                       N/S-5845520129 '},
        {'id': 34, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-Tetra Centri MRPX318TGV             N/S-2969461 '},
        {'id': 35, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-Tetra Centri MRPX318TGV             N/S-2969463 '},
        {'id': 36, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-Tetra Centri MRPX318TGV             N/S-2969464 '},
        {'id': 37, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-Tetra Plex MS10 SBL                 N/S-30104-38510 '},
        {'id': 38, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-Tetra Plex MS10 SBL                 N/S-30104-38511 '},
        {'id': 39, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-Tetra Plex MS6 SBL                  N/S-30104-38509 '},      
        {'id': 40, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-Tetra Plex Plex C8                  N/S-30104-38508 '},
        {'id': 41, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-Tetra Plex, C8                      N/S-30107-34617 '},
        {'id': 42, 'nombre': 'UYP-CONAPROLE CIM-UHT01   Máquina-Tetra Pak Aseptic Tank LV           N/S-T5845520044 '},
        {'id': 43, 'nombre': 'UYP-CEREALIN  SJO-TREAT   Máquina-TETRA ALEX 20                       N/S-5845511263'},
        {'id': 44, 'nombre': 'UYP-CEREALIN  SJO-TREAT   Máquina-Tetra Therm Aseptic Flex 1          N/S-T5844410001'},
        {'id': 45, 'nombre': 'UYP-BARRACA   MVD-PF 01   Máquina-Tetra Alex 25                       N/S-T5856826141'},
        {'id': 46, 'nombre': 'UYP-BARRACA   MVD-PF 01   Máquina-Tetra Pak Homogenizer               N/S-5856944267'},
        {'id': 47, 'nombre': 'UYP-CONAPROLE P01-TREAT   Máquina-Tetra Plex Clip 6 RM                 N/S-30103-22150'},
        {'id': 48, 'nombre': 'UYP-CONAPROLE P01-TPPM    Máquina-Production Integrator                N/S-5861001701'},
        {'id': 49, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Tetra Pak® Separator A2              N/S-C200001'},
        {'id': 50, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Separator A2                         N/S-4269765'},
        {'id': 51, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Tetra Pak® Separator A2              N/S-C200002'},
        {'id': 52, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Separator A2                         N/S-4269764'},
        {'id': 53, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Tetra Pak® Separator H10             N/S-C200003'},
        {'id': 54, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Separator H10                        N/S-4269763'},
        {'id': 55, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Tetra Pak® Separator H714            N/S-C200004'},
        {'id': 56, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Separator H714                       N/S-4260456'},
        {'id': 57, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Tetra Pak Homogenizer                N/S-5870520020'},
        {'id': 58, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-HWU                                  N/S-D6295377/04'},
        {'id': 59, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Drum Filler                          N/S-PR2438'},
        {'id': 60, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-PHE cip                              N/S-30105-09114'},
        {'id': 61, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Skid Crema                           N/S-D6295377/01'},
        {'id': 62, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-PHE Crema                            N/S-30125-10251'},
        {'id': 63, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Skid Polisher + Desaireador          N/S-D6295377/02'},
        {'id': 64, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Tetra Plex C6-SM - AMF               N/S-30125-10250'},
        {'id': 65, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Tanque Pulmon                        N/S-D6295377/03'},
        {'id': 66, 'nombre': 'UYP-CONAPROLE P07-CH 01   Máquina-Tetra Pak PHE, M10                   N/S-30125-12013'},
        {'id': 67, 'nombre': 'UYP-CONAPROLE P07-TREAT   Máquina-Tetra Plex Clip 10 RH                N/S-30103-17680'},
        {'id': 68, 'nombre': 'UYP-CONAPROLE P07-TREAT   Máquina-Tetra Plex Clip 6 RM                 N/S-30103-17688'},
        {'id': 70, 'nombre': 'UYP-CONAPROLE P07-TREAT   Máquina-Tetra Centri HMRPX718HGV74C          N/S-4110484'},
        {'id': 71, 'nombre': 'UYP-CONAPROLE P07-TREAT   Máquina-Tetra Centri HMRPX718HGV74C          N/S-4110489'},
        {'id': 72, 'nombre': 'UYP-CONAPROLE P07-TREAT   Máquina-Tetra Plex C6                        N/S-30104-21165'},
        {'id': 73, 'nombre': 'UYP-CONAPROLE P07-TREAT   Máquina-Tetra Plex C8                        N/S-30104-21168'},
        {'id': 74, 'nombre': 'UYP-CONAPROLE P07-TREAT   Máquina-Tetra Plex M10 M BASE                N/S-30103-17682'},
        {'id': 75, 'nombre': 'UYP-CONAPROLE P07-TREAT   Máquina-Tetra Plex Clip 10 RM                N/S-30103-17681'},
        {'id': 76, 'nombre': 'UYP-CONAPROLE P08-DR 01   Máquina-Pasteurizador 4ta línea estandarización       N/S-D6333928/01'},
        {'id': 77, 'nombre': 'UYP-CONAPROLE P08-DR 01   Máquina-Tetra Pak PHE, C15 Sanitary                   N/S-30125-35464'},
        {'id': 78, 'nombre': 'UYP-CONAPROLE P08-DR 01   Máquina-Tetra Pak PHE, M6                             N/S-30125-11940'},
        {'id': 79, 'nombre': 'UYP-CONAPROLE P08-DR 01   Máquina-Tetra Pak Standardization unit                N/S-T5845470166'},
        {'id': 80, 'nombre': 'UYP-CONAPROLE P08-DR 01   Máquina-Tetra Pak® Separator H80                      N/S-C220009'},
        {'id': 81, 'nombre': 'UYP-CONAPROLE P08-DR 01   Máquina-Separator H80                                  N/S-AAF0000430'},
        {'id': 82, 'nombre': 'UYP-CONAPROLE P08-MX 01   Máquina-Tetra Pak High Shear Mixer                     N/S-T5845706418'},
        {'id': 83, 'nombre': 'UYP-CONAPROLE P08-MX 01   Máquina-Tetra Pak High Shear Mixer                     N/S-T5845706420'},
        {'id': 84, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Powder Handling                      N/S-5860700023'},
        {'id': 85, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Big Bag Tipping unit VB005          N/S-5860010176'},
        {'id': 86, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Big Bag Tipping unit VB005          N/S-5860010177'},
        {'id': 87, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040608'},
        {'id': 88, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040609'},
        {'id': 89, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Lump Breaker TD041                  N/S-5860190027'},
        {'id': 90, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Bag Tipping unit M VS401           N/S-5860020148'},
        {'id': 91, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Dust Filter C FL002                 N/S-5860170368'},
        {'id': 92, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Rotary Valve FB                     N/S-5860070145'},
        {'id': 93, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Rotary Valve FB                     N/S-5860070146'},
        {'id': 94, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®  Inlet Air Managment PLP          N/S-5860240325'},
        {'id': 95, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®  Outlet Air Managment PLP         N/S-5860240326'},
        {'id': 96, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280654'},
        {'id': 97, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040610'},
        {'id': 98, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040611'},
        {'id': 99, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280655'},
        {'id': 100, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280656'},
        {'id': 101, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280657'},
        {'id': 102, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280658'},
        {'id': 103, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040612'},
        {'id': 104, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040613'},
        {'id': 105, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040614'},
        {'id': 106, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040615'},
        {'id': 107, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040616'},
        {'id': 108, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®  Inlet Air Managment PLP          N/S-5860240327'},
        {'id': 109, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®  Outlet Air Managment PLP         N/S-5860240328'},
        {'id': 110, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Rotary Valve FB                    N/S-5860070147'},
        {'id': 111, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Rotary Valve FB                    N/S-5860070148'},
        {'id': 112, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Dust Filter C FL002                N/S-5860170369'},
        {'id': 113, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Dust Filter C FL002                N/S-5860170370'},
        {'id': 114, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280659'},
        {'id': 115, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280660'},
        {'id': 116, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280661'},
        {'id': 117, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280662'},
        {'id': 118, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Big Bag Filling unit S DB006       N/S-5860140027'},
        {'id': 119, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Bag Tipping unit CS VS600          N/S-5860310077'},
        {'id': 120, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Bag Tipping unit CS VS600          N/S-5860310078'},
        {'id': 121, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Bag Tipping unit CS VS600          N/S-5860310079'},
        {'id': 122, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Bag Tipping unit CS VS600          N/S-5860310080'},
        {'id': 123, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280663'},
        {'id': 124, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280664'},
        {'id': 125, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280665'},
        {'id': 126, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280666'},
        {'id': 127, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Lump Breaker TD041                 N/S-5860190030'},
        {'id': 128, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                  N/S-5860040689'},
        {'id': 129, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Centri MRPX214TGV44C                   N/S-2917438'},
        {'id': 130, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Centri MRPX214TGV44                    N/S-2985490'},
        {'id': 131, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Centri MRPX214-74T                     N/S-2893664'},
        {'id': 132, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Centri MRPX214-TGV74                   N/S-2898316'},
        {'id': 133, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Alfast 220                             N/S-T5845420205'},
        {'id': 134, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Plex Clip 10 RM                        N/S-30103-22146'},
        {'id': 135, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Therm Lacta                            N/S-D00623-01/1'},
        {'id': 136, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Plex M10 B                             N/S-30100-22169'},
        {'id': 137, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Plex M6 M                              N/S-30103-22147'},
        {'id': 139, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Plex M6 MFMC                           N/S-30100-41336'},
        {'id': 140, 'nombre': 'UYP-CONAPROLE P08-UHT01   Máquina-Tetra Pak Aseptic Tank LV                    N/S-T5845520045'},
        {'id': 141, 'nombre': 'UYP-CONAPROLE P08-UHT01   Máquina-Tank body Tetra Pak Aseptic Tank VD/LV       N/S-211182'},
        {'id': 142, 'nombre': 'UYP-CONAPROLE P08-UHT01   Máquina-Tetra Pak Aseptic Dosing unit F              N/S-T5844550349'},
        {'id': 143, 'nombre': 'UYP-CONAPROLE P08-TPPM    Máquina-Production Control                           N/S-5861001994'},
        {'id': 144, 'nombre': 'UYP-CONAPROLE P08-MP 01   Máquina-Tetra Therm Lacta                            N/S-T5845140314'},
        {'id': 145, 'nombre': 'UYP-CONAPROLE P08-MP 01   Máquina-Tetra Plex, C10                              N/S-30109-16965'},
        {'id': 146, 'nombre': 'UYP-CONAPROLE P08-MP 01   Máquina-Tetra Pak Homogenizer                        N/S-5871117115'},
        {'id': 147, 'nombre': 'UYP-CONAPROLE P09-CH 01   Máquina-Tetra Tebel Ost IV                           N/S-620910.01'},
        {'id': 148, 'nombre': 'UYP-CONAPROLE P09-CH 01   Máquina-Tetra Tebel Ost IV                           N/S-620910.02'},
        {'id': 149, 'nombre': 'UYP-CONAPROLE P09-CH 01   Máquina-Tetra Tebel Ost IV                           N/S-620910.03'},
        {'id': 150, 'nombre': 'UYP-CONAPROLE P09-CH 01   Máquina-Tetra Tebel Ost IV                           N/S-620910.04'},
        {'id': 151, 'nombre': 'UYP-CONAPROLE P09-CH 01   Máquina-Tetra Tebel Alfomatic 2B                     N/S-620922.01'},
        {'id': 152, 'nombre': 'UYP-CONAPROLE P09-CH 01   Máquina-Tetra Tebel OST SH 5.1 12.500 L              N/S-5847110077'},
        {'id': 153, 'nombre': 'UYP-CONAPROLE P09-CH 01   Máquina-Tetra Damrow Double OO CH 8.0 14.000L        N/S-5847190007'},
        {'id': 154, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Alfast 220                             N/S-T5845420207'},
        {'id': 155, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Alfast 220                             N/S-T5845420206'},
        {'id': 156, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Plex Clip 10 RM                        N/S-30103-22144'},
        {'id': 157, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Centri HMRPX414HGV74C                  N/S-417117'},
        {'id': 159, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Centri HMRPX614HGV74C                  N/S-4042654'},
        {'id': 160, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Therm Lacta                            N/S-D00623-01/3'},
        {'id': 161, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Plex M6 M                              N/S-30103-22149'},
        {'id': 162, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Plex C8-KSR                            N/S-30106-34564'},
        {'id': 163, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Plex, C8                               N/S-30109-15185'},
        {'id': 164, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Plex, M10                              N/S-30109-15186'},
        {'id': 165, 'nombre': 'UYP-CONAPROLE P11-TREAT   Máquina-Tetra Centri HMRPX514HGV74C                  N/S-4033885'},
        {'id': 166, 'nombre': 'UYP-CONAPROLE P11-TREAT   Máquina-Tetra Centri HMRPX714HGV74C                  N/S-4047653'},
        {'id': 167, 'nombre': 'UYP-CONAPROLE P16-TREAT   Máquina-Tetra Plex Clip 10 RM                        N/S-30103-22145'},
        {'id': 168, 'nombre': 'UYP-CONAPROLE P16-TREAT   Máquina-Tetra Therm Lacta                            N/S-D00623-01/2'},
        {'id': 169, 'nombre': 'UYP-CONAPROLE P16-TREAT   Máquina-Tetra Plex M6 M                              N/S-30103-22148'},
        {'id': 170, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Contisolv                         	  N/S-15735-01'},
        {'id': 171, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco CIP unit                          	  N/S-15735-02'},
        {'id': 172, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco CIP unit                          	  N/S-15735-03'},
        {'id': 173, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Thermizer                         	  N/S-15735-04'},
        {'id': 174, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco CIP unit                          	  N/S-15735-05'},
        {'id': 175, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Ion Exchange System                   N/S-15735-06'},
        {'id': 176, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Horizontal Plate Filtration           N/S-16676-01'},
        {'id': 177, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco RJ Mixer                              N/S-16958-02'},
        {'id': 178, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco RJ Mixer                              N/S-16958-03'},
        {'id': 179, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco RJ Mixer                              N/S-16958-04'},
        {'id': 180, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco RJ Mixer                              N/S-16958-05'},
        {'id': 131, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Ingredients Dissolver                 N/S-15229-01'},
        {'id': 182, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Filtration                            N/S-15735-07'},
        {'id': 183, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Filtration                            N/S-15735-08'},
        {'id': 184, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Ingredients Dissolver                 N/S-16676-'},
        {'id': 185, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Ingredients Dissolver                 N/S-T5880160029'},
        {'id': 186, 'nombre': 'UYP-MONRESA   MVD-MX01    Máquina-Tetra Pak Mixer RJCI                         N/S-T5845680002'},
        {'id': 187, 'nombre': 'UYC-JUGOSURUG FRY-LIN01   Máquina- TBA/21                                      N/S-15011/00099'},
        {'id': 188, 'nombre': 'UYC-JUGOSURUG FRY-LIN02   Máquina- A3/Flex                                     N/S-21211/00147'},
        {'id': 189, 'nombre': 'UYC-JUGOSURUG FRY-LIN03   Máquina- A3/Flex                                     N/S-21211/00389'},
        {'id': 190, 'nombre': 'UYP-JUGOSURUG FRY-TREAT   Máquina-Tetra Plex Clip 6 RM WS                      N/S-30101-22606'},
        {'id': 191, 'nombre': 'UYP-JUGOSURUG FRY-TREAT   Máquina-Tetra Plex CLIP 6 WS                         N/S-30101-25950'},
        {'id': 192, 'nombre': 'UYP-JUGOSURUG FRY-TREAT   Máquina-Tetra Therm Nacional                         N/S-95085'},
        {'id': 193, 'nombre': 'UYP-JUGOSURUG FRY-TREAT   Máquina-Tetra Therm Nacional                         N/S-B97439-01'},
        {'id': 194, 'nombre': 'UYP-JUGOSURUG FRY-Z_DOC   Máquina-Plant Documentation                          N/S-D5865000135'},
        {'id': 195, 'nombre': 'UYC-CONAPROLE P08-LIN05   Máquina-E3/Speed                                     N/S-21225/00033'},
        {'id': 196, 'nombre': 'UYP-ECOLAT    NVH-TREAT TETRA ALEX 30 A   Máquina-T4625220                     N/S-T4625220'},
        {'id': 197, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Alsafe LA   Máquina-T2440170                     N/S-T2440170'},
        {'id': 198, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Therm Aseptic Flex   Máquina-T2100197            N/S-T2100197'},
        {'id': 199, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Alfast Plus   Máquina-T5845440334                N/S-T5845440334'},
        {'id': 200, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Plex MS6-SR in C6092071/04   Máquina-30106-35480 N/S-30106-35480'},
        {'id': 201, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Plex C6-SR in C6092071/01   Máquina-30106-35479  N/S-30106-35479'},
        {'id': 202, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Alcip Nacional   Máquina-C6092071/02             N/S-C6092071/02'},
        {'id': 203, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Therm Nacional   Máquina-C6092071/04             N/S-C6092071/04'},
        {'id': 204, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Therm Nacional   Máquina-C6092071/01             N/S-C6092071/01'},
        {'id': 205, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Therm Nacional   Máquina-C6092071/03             N/S-C6092071/03'},
        {'id': 206, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Plex MS6-SR in C6092071/03   Máquina-30106-35573 N/S-30106-35573'},
        {'id': 207, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Plex C6-SR   Máquina-30106-35478                 N/S-30106-35478'},
        {'id': 208, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Plex C6-SR   Máquina-30106-35477                 N/S-30106-35477'},
        {'id': 209, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Alcross US 100   Máquina-T5845810090             N/S-T5845810090'},
        {'id': 210, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Plex MS10-SR   Máquina-30107-34354               N/S-30107-34354'},
        {'id': 211, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Alfast   Máquina-T5845440556                     N/S-T5845440556'},
        {'id': 212, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Alfast   Máquina-T5845440557                     N/S-T5845440557'},
        {'id': 213, 'nombre': 'UYP - N/A'},
        {'id': 214, 'nombre': 'UYC - N/A'}
    ]



    if request.method == 'POST':
        fecha = request.form['fecha']
        entrada = f"{request.form['entrada_hora']}:{request.form['entrada_minuto']}"
        salida = f"{request.form['salida_hora']}:{request.form['salida_minuto']}"


        try:
            almuerzo_horas = int(request.form.get('almuerzo_horas', 0))
            almuerzo_minutos = int(request.form.get('almuerzo_minutos', 0))
        except ValueError:
            flash("El tiempo de almuerzo debe ser un número válido", "danger")
            return redirect(url_for('dashboard'))

        almuerzo = timedelta(hours=almuerzo_horas, minutes=almuerzo_minutos)

        try:
            viaje_ida = float(request.form.get('viaje_ida', 0) or 0)
            viaje_vuelta = float(request.form.get('viaje_vuelta', 0) or 0)
            km_ida = float(request.form.get('km_ida', 0) or 0)
            km_vuelta = float(request.form.get('km_vuelta', 0) or 0)
        except ValueError:
            flash("Las horas de viaje y kilómetros deben ser números válidos.", "danger")
            return redirect(url_for('dashboard'))

        tarea = request.form.get('tarea', '').strip()
        cliente = request.form.get('cliente', '').strip()  # Si cliente es texto, OK
        comentarios = request.form.get('comentarios', '').strip()
        #contrato = bool(int(request.form.get("contrato")))
        contrato = request.form.get("contrato", "").strip()

        service_order = request.form.get('service_order', '').strip()

        try:
            centro_costo = request.form.get('centro_costo', '').strip()
            tipo_servicio = request.form.get('tipo_servicio', '').strip()
            linea = request.form.get('linea', '').strip()

        except ValueError:
            flash("Los campos de selección deben ser valores válidos.", "danger")
            return redirect(url_for('dashboard'))

        try:
            formato_hora = "%H:%M"
            t_entrada = datetime.strptime(entrada, formato_hora)
            t_salida = datetime.strptime(salida, formato_hora)
            if t_salida < t_entrada:
                t_salida += timedelta(days=1)
            tiempo_total = t_salida - t_entrada - almuerzo
            horas_trabajadas = tiempo_total.total_seconds() / 3600
        except ValueError:
            flash("Formato de hora incorrecto. Use HH:MM.", "danger")
            return redirect(url_for('dashboard'))
            

        nuevo_registro = Registro(
            user_id=session['user_id'],
            fecha=fecha,
            entrada=entrada,
            salida=salida,
            almuerzo=round(almuerzo.total_seconds() / 3600, 2),
            horas=round(horas_trabajadas, 2),
            viaje_ida=viaje_ida,
            viaje_vuelta=viaje_vuelta,
            km_ida=km_ida,
            km_vuelta=km_vuelta,
            tarea=tarea,
            cliente=cliente,
            comentarios=comentarios,
            contrato=contrato,
            service_order=service_order,
            centro_costo=centro_costo,
            tipo_servicio=tipo_servicio,
            linea=linea
        )

        db.session.add(nuevo_registro)
        db.session.commit()
        flash('Registro guardado exitosamente', category='success')
        return redirect(url_for('dashboard'))

    # GET - mostrar registros y total de horas
    filtros = request.args
    registros_query = Registro.query.filter_by(user_id=session['user_id'])
    if 'fecha' in filtros:
        registros_query = registros_query.filter_by(fecha=filtros['fecha'])
    registros = registros_query.order_by(Registro.fecha.desc()).all()

    total_horas = sum([(r.horas or 0) + (r.viaje_ida or 0) + (r.viaje_vuelta or 0) for r in registros])
    total_km = sum([(r.km_ida or 0) + (r.km_vuelta or 0) for r in registros])
      
      
    cliente_prefijo = {
        '1561046 - Barraca Deambrosi SA (Carton)'   : 'UYC-BARRACA',
        '1561046 - Barraca Deambrosi SA (Proceso)'  : 'UYP-BARRACA',
        '1561458 - Belficor Sociedad Anónima (Proceso)'       : 'UYP-BELFICOR',
        '1561459 - Belficor - Paraje La Boyada'     : 'UYP-BELFICOR',
        '1561233 - Companía Salus SA'               : 'UYP-FTE SALUS',
        '1561251 - Cooperativa Agraria (CALCAR)'    : 'UYC-COAGRARIA',
        '1560621 - Ecolat Uruguay S.A. (Proceso)'   : 'UYP-ECOLAT',
        '1561335 - Gibur S.A. (Carton)'             : 'UYC-GIBUR',
        '1561540 - Nolir S.A.'                      : 'UYC-NOLIR',
        '1560020 - Recalco S.A. (ex Suadil)'        : 'UYC-RECALCO',
        '1561245 - Conaprole Planta I (Proceso)'    : 'UYP-CONAPROLE P01',
        '1561250 - Conaprole Planta CIM (Carton)'   : 'UYC-CONAPROLE CIM',
        '1561250 - Conaprole Planta CIM (Proceso)'  : 'UYP-CONAPROLE CIM',
        '1561010 - Conaprole Planta VII (Proceso)'  : 'UYP-CONAPROLE P07',
        '1560606 - Conaprole Planta VIII (Carton)'  : 'UYC-CONAPROLE P08',
        '1560606 - Conaprole Planta VIII (Proceso)' : 'UYP-CONAPROLE P08',
        '1560018 - Cerealin San Jose (Carton)'      : 'UYC-CEREALIN',
        '1560018 - Cerealin San Jose (Proceso)'     : 'UYP-CEREALIN',
        '1569004 - Cerealin S.A.'          : 'UYP-CEREALIN',
        '1561132 - Montevideo Refrescos S.A. (Proceso)':'UYP-MONRESA',
        '1561062 - Jugos del Uruguay S.A. (Carton)'   : 'UYC-JUGOSURUG',  # definir si hay prefijo
        '1561062 - Jugos del Uruguay S.A. (Proceso)'  : 'UYP-JUGOSURUG',  # definir si hay prefijo
        '1570011 - Tetrapak Argentina'           : 'N/A',
        'Otros Clientes del Mercado'      : 'UYP - N/A',
        'Otros Clientes del Mercado'      : 'UYC - N/A'
    }
      # ─── Construcción automática de cliente_cc_lineas ───
    
    cliente_cc_lineas = {}
    for cli in clientes:
        centros = [cc['nombre'] for cc in centros_costo if cli in cc['nombre']]
        pref = cliente_prefijo.get(cli, '')
        if pref:
            lineas_f = [ln['nombre'] for ln in lineas if ln['nombre'].startswith(pref)]
        else:
            lineas_f = []
        cliente_cc_lineas[cli] = {
            'centros_costo': centros,
            'lineas':        lineas_f
        }
    contrato_labels = {item['value']: item['label'] for item in contratos}

    return render_template(
        'dashboard.html',
        username=session['username'],
        role=session['role'],
        registros=registros,
        total_horas=round(total_horas, 2),
        total_km=round(total_km, 2),
        clientes=clientes,
        contratos=contratos,
        contrato_labels=contrato_labels,
        service_orders=service_orders,
        centros_costo=centros_costo,
        tipos_servicio=tipos_servicio,
        lineas=lineas,
        cliente_cc_lineas = cliente_cc_lineas,
        tipo_cliente=tipo_cliente
    )


@app.route('/exportar_excel')
def exportar_excel():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    role = session.get('role')
    user_id = session.get('user_id')
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    contexto = request.args.get('contexto')
    usuario_id = request.args.get('usuario_id')

    query = Registro.query

    if role in ['admin', 'superadmin'] and contexto != 'admin':
        query = query.filter_by(user_id=user_id)
    elif role not in ['admin', 'superadmin']:
        query = query.filter_by(user_id=user_id)

    if usuario_id and role in ['admin', 'superadmin']:
        query = query.filter_by(user_id=usuario_id)

    if not fecha_desde or not fecha_hasta or fecha_desde == 'None' or fecha_hasta == 'None':
        flash("Debés completar ambas fechas para exportar.", "warning")
        return redirect(url_for('admin'))


    query = query.filter(Registro.fecha.between(fecha_desde, fecha_hasta))
    registros = query.all()

    df = pd.DataFrame([{
        'Usuario': r.user.username if r.user else None,
        'Fecha': r.fecha,
        'Entrada': r.entrada,
        'Salida': r.salida,
        'Almuerzo (hs)': r.almuerzo,
        'Viaje ida (hs)': r.viaje_ida,
        'Viaje vuelta (hs)': r.viaje_vuelta,
        'Horas laborales': r.horas,
        'Horas totales': round((r.horas or 0) + (r.viaje_ida or 0) + (r.viaje_vuelta or 0), 2),
        'Km ida': r.km_ida,
        'Km vuelta': r.km_vuelta,
        'Km totales': (r.km_ida or 0) + (r.km_vuelta or 0),
        'Tarea': r.tarea,
        'Cliente': r.cliente,
        'Comentarios': r.comentarios,
        #'Contrato': 'Sí' if r.contrato else 'N/A',
        'Contable': r.contrato if r.contrato else 'N/A',
        'Service Order': r.service_order or '',
        'Centro de Costo': r.centro_costo or '',
        'Tipo de Servicio': r.tipo_servicio or '',
        'Línea': r.linea or ''
    } for r in registros if r.user is not None])

    df['Contable'] = df['Contable'].map({
        '73450003': 'Contrato',
        '79010000': 'Administrativo - 79010000',
        '79200030': 'Issue Resolution - Service - 79200030',
        '79200020': 'Issue Resolution - Capital CARTON - 79200020',
        '79200050': 'Issue Resolution - Capital PROCESSING - 79200050',
        '580000': 'Capital Equipment carton - 580000',
        '70350000': 'Capital Equipment processing - 70350000',
        '73450000': 'Fuera de contrato - 73450000',
        '560000': 'Upgrade projects - 560000',
        '70158000': 'Mandatory kit - 70158000',
        '480000': 'Rebilling - 480000',
        '73453000': 'Training Facturable - 73453000',
        '70400010': 'MDR - 70400010',
        '0': 'N/A'
    }).fillna(df['Contable'])


    archivo = BytesIO()
    with pd.ExcelWriter(archivo, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Registros', startrow=0)
        ws = writer.sheets['Registros']

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", name='Calibri')
        header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Encabezados
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_alignment

        # Celdas
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font = Font(name='Calibri', size=11)
                cell.border = thin_border
                cell.alignment = center_alignment
            if row[0].row % 2 == 1:
                for cell in row:
                    cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

        # Ajuste de columnas
        for col_num, column_cells in enumerate(ws.columns, 1):
            max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0)
            adjusted_width = min((max_length + 4), 50)
            ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width

        # Agregar fila de totales
        total_row = ws.max_row + 2
        ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)

        for col in ws.iter_cols(min_row=1, max_row=1):
            header = col[0].value
            if header == "Horas laborales":
                col_idx = col[0].column
                total = df["Horas laborales"].sum()
                cell = ws.cell(row=total_row, column=col_idx, value=round(total, 2))
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.border = thin_border
                cell.alignment = center_alignment
            elif header == "Km totales":
                col_idx = col[0].column
                total = df["Km totales"].sum()
                cell = ws.cell(row=total_row, column=col_idx, value=round(total, 2))
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.border = thin_border
                cell.alignment = center_alignment

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'

    archivo.seek(0)
    return send_file(
        archivo,
        as_attachment=True,
        download_name=f"registros_{session['username']}.xlsx"
    )





@app.route('/editar_registro/<int:id>', methods=['GET', 'POST'])
def editar_registro(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Capturar filtros desde la URL

    filtro_usuario = request.args.get('filtro_usuario')
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    

    registro = Registro.query.get_or_404(id)

    # Definí acá tus listas o importalas desde donde estén definidas
    
    clientes = [
        '1561046 - Barraca Deambrosi SA (Carton)',
        '1561046 - Barraca Deambrosi SA (Proceso)',
        '1561458 - Belficor Sociedad Anónima (Proceso)',
        '1561459 - Belficor - Paraje La Boyada',
        '1560018 - Cerealin San Jose (Carton)',
        '1560018 - Cerealin San Jose (Proceso)',
        '1569004 - Cerealin S.A.',
        '1561233 - Companía Salus SA',
        '1561245 - Conaprole Planta I (Proceso)',
        '1561250 - Conaprole Planta CIM (Carton)',
        '1561250 - Conaprole Planta CIM (Proceso)',
        '1561010 - Conaprole Planta VII (Proceso)',
        '1560606 - Conaprole Planta VIII (Carton)',
        '1560606 - Conaprole Planta VIII (Proceso)',
        '1561251 - Cooperativa Agraria (CALCAR)',
        '1560621 - Ecolat Uruguay S.A. (Proceso)',
        '1561335 - Gibur S.A. (Carton)',
        '1561062 - Jugos del Uruguay S.A. (Carton)',
        '1561062 - Jugos del Uruguay S.A. (Proceso)',
        '1561132 - Montevideo Refrescos S.A. (Proceso)',
        '1561540 - Nolir S.A.',
        '1560020 - Recalco S.A. (ex Suadil)',
        '1570011 - Tetrapak Argentina',
        'Otros Clientes del Mercado'
    ]


    #contratos = ['Contrato legal 1', 'Contrato legal 2', 'Contrato legal 3']

    service_orders = ['SM02', 'SM03', 'N/A']

    centros_costo = [
        {'id': 1, 'nombre': '1561046 - Barraca Deambrosi SA (Carton) C.Costo=1 ()'},
        {'id': 2, 'nombre': '1561251 - Cooperativa Agraria (CALCAR) C.Costo=2 ()'},
        {'id': 3, 'nombre': '1561335 - Gibur S.A. (Carton) C.Costo=3 ()'},
        {'id': 4, 'nombre': '1561540 - Nolir S.A. C.Costo=4 ()'},
        {'id': 5, 'nombre': '1560020 - Recalco S.A. (ex Suadil) C.Costo=5 ()'},
        {'id': 6, 'nombre': '1561250 - Conaprole Planta CIM (Carton) C.Costo=6 ()'},
        {'id': 7, 'nombre': '1561245 - Conaprole Planta I (Proceso) C.Costo= ()'},
        {'id': 8, 'nombre': '1561250 - Conaprole Planta CIM (Proceso) C.Costo= ()'},
        {'id': 9, 'nombre': '1561010 - Conaprole Planta VII (Proceso) C.Costo= ()'},
        {'id': 10, 'nombre': '1560606 - Conaprole Planta VIII (Carton) C.Costo=7 ()'},
        {'id': 11, 'nombre': '1560606 - Conaprole Planta VIII (Proceso) C.Costo= ()'},
        {'id': 12, 'nombre': '1560018 - Cerealin San Jose (Carton)C.Costo=8 ()'},
        {'id': 13, 'nombre': '1560018 - Cerealin San Jose (Proceso) C.Costo=8 ()'},
        {'id': 14, 'nombre': '1561062 - Jugos del Uruguay S.A. (Carton) GMB revisar ()'},
        {'id': 15, 'nombre': '1561062 - Jugos del Uruguay S.A. (Proceso) GMB revisar ()'},
        {'id': 16, 'nombre': 'FUERA DE CONTRATO'},       
        {'id': 17, 'nombre': '1561458 - Belficor Sociedad Anónima (Proceso) C.Costo='},
        {'id': 18, 'nombre': '1561459 - Belficor - Paraje La Boyada C.Costo='},
        {'id': 19, 'nombre': '1561233 - Companía Salus SA C.Costo='},
        {'id': 20, 'nombre': 'Barraca Deambrosi SA (Proceso) C.Costo=1 ()'},
        {'id': 22, 'nombre': '1561132 - Montevideo Refrescos S.A. (Proceso) C.Costo= ()'},
        {'id': 23, 'nombre': '1569004 - Cerealin S.A. C.Costo= ()'},
        {'id': 24, 'nombre': '1560621 - Ecolat Uruguay S.A. (Proceso) C.Costo= ()'},
        {'id': 25, 'nombre': 'Otros Clientes del Mercado N/A'}
    ]
    contratos = [
            {'value': '73450003', 'label': 'Contrato'},
            {'value': '79010000', 'label': 'Administrativo - 79010000'},
            {'value': '79200030', 'label': 'Issue Resolution - Service - 79200030'},
            {'value': '79200020', 'label': 'Issue Resolution - Capital CARTON - 79200020'},
            {'value': '79200050', 'label': 'Issue Resolution - Capital PROCESSING - 79200050'},
            {'value': '580000', 'label': 'Capital Equipment carton - 580000'},
            {'value': '70350000', 'label': 'Capital Equipment processing - 70350000'},
            {'value': '73450000', 'label': 'Fuera de contrato - 73450000'},
            {'value': '560000', 'label': 'Upgrade projects - 560000'},
            {'value': '70158000', 'label': 'Mandatory kit - 70158000'},
            {'value': '480000', 'label': 'Rebilling - 480000'},
            {'value': '73453000', 'label': 'Training Facturable - 73453000'},
            {'value': '79010000', 'label': 'Training Interno - 79010000'},
            {'value': '70400010', 'label': 'MDR - 70400010'},
            {'value': '79010000', 'label': 'No disponible - 79010000'},
            {'value': '0', 'label': 'N/A'}
    ]

    tipos_servicio = [
        {'id': 1, 'nombre': 'TPMS'},
        {'id': 2, 'nombre': 'Correctivo'},
        {'id': 3, 'nombre': 'Soporte a la producción'},
        {'id': 4, 'nombre': 'Proyectos'},
        {'id': 5, 'nombre': 'RK'},
        {'id': 6, 'nombre': 'Vacaciones'},
        {'id': 7, 'nombre': 'Tareas Administrativas'},
        {'id': 8, 'nombre': 'Compensatorios'},
        {'id': 9, 'nombre': 'Training'},
        {'id': 10, 'nombre': 'Claims'}
    ]

    lineas = [
        {'id': 1,  'nombre': 'UYC-BARRACA   MVD-LIN01   Máquina-TBA/3                               N/S-11443/05537'},
        {'id': 2,  'nombre': 'UYC-BARRACA   MVD-LIN02   Máquina-TBA/8                               N/S-20201/82004'},
        {'id': 3,  'nombre': 'UYC-BARRACA   MVD-LIN03   Máquina-SIMPLY8                             N/S-21222/00018'},
        {'id': 4,  'nombre': 'UYC-BARRACA   MVD-LIN04   Máquina-TBA/19                              N/S-20562/83308'},
        {'id': 5,  'nombre': 'UYC-COAGRARIA CAR-LN 01   Máquina-TBA/8                               N/S-13037/10830'},
        {'id': 6,  'nombre': 'UYC-COAGRARIA CAR-LN 02   Máquina-TP C3/F                             N/S-15034/00004'},
        {'id': 7,  'nombre': 'UYC-NOLIR     MVD-LIN01   Máquina-TBA/19                              N/S-20591/83337'},
        {'id': 8,  'nombre': 'UYC-NOLIR     MVD-LIN02   Máquina-TBA/8                               N/S-15010/00889'},
        {'id': 9,  'nombre': 'UYC-CEREALIN  SJO-LIN01   Máquina-TBA/8                               N/S-13588/11417'},
        {'id': 10, 'nombre': 'UYC-CEREALIN  SJO-LIN04   Máquina-TP A3/CF                            N/S-21220/00466'},
        {'id': 11, 'nombre': 'UYC-CONAPROLE CIM-LIN02   Máquina-TBA/19                              N/S-20258/82571'},
        {'id': 12, 'nombre': 'UYC-CONAPROLE CIM-LIN03   Máquina-TT/3                                N/S-63202/20090'},
        {'id': 13, 'nombre': 'UYC-CONAPROLE P08-LIN01   Máquina-TBA/8                               N/S-20239/82382'},
        {'id': 14, 'nombre': 'UYC-CONAPROLE P08-LIN02   Máquina-TBA/8                               N/S-13879/11665'},
        {'id': 15, 'nombre': 'UYC-CONAPROLE P08-LIN03   Máquina-TBA/8                               N/S-13457/11304'},
        {'id': 16, 'nombre': 'UYC-CONAPROLE P08-LIN04   Máquina-TBA/8                               N/S-13486/11332'},
        {'id': 17, 'nombre': 'UYC-GIBUR     MVD-LIN01   Máquina-TBA/8                               N/S-17010/00018'},
        {'id': 18, 'nombre': 'UYC-GIBUR     MVD-LIN02   Máquina-TBA/19                               N/S-21200/00201'},
        {'id': 19, 'nombre': 'UYC-RECALCO   MVD-LIN01   Máquina-TBA/3                               N/S-20078/80780'},
        {'id': 20, 'nombre': 'UYC-RECALCO   MVD-LIN02   Máquina-TBA/8                               N/S-12967/10664'},
        {'id': 21, 'nombre': 'UYP-BELFICOR  MDV-CIP     Máquina-Tetra Alcip 2                       N/S-C6075349/01'},
        {'id': 22, 'nombre': 'UYP-FTE SALUS MIN-TREAT   Máquina-Tetra Therm Nacional                N/S-B04013-02'},
        {'id': 23, 'nombre': 'UYP-FTE SALUS MIN-TREAT   Máquina-Tetra Plex MS6 SR                   N/S-30105-08295'},
        {'id': 24, 'nombre': 'UYP-FTE SALUS MIN-TREAT   Máquina-Tetra Plex C8 in B6069919/01        N/S-30106-34366'},
        {'id': 25, 'nombre': 'UYP-FTE SALUS MIN-TREAT   Máquina-Tetra Therm Nacional                N/S-B6069919/01'},
        {'id': 26, 'nombre': 'UYP-FTE SALUS MIN-TREAT   Máquina-Tetra Alcip Nacional                N/S-B6069919/02'},
        {'id': 27, 'nombre': 'UYP-FTE SALUS MIN-TREAT   Máquina-Tanque Ultra Clean Nacional         N/S-B6123596/01'},
        {'id': 28, 'nombre': 'UYP-FTE SALUS MIN-TREAT   Máquina-Modulo CIP                          N/S-B6249678/01 '},
        {'id': 29, 'nombre': 'UYP-CONAPROLE CIM-IC 01   Máquina-Tetra Pak® Ingredient Doser 4000 A2 N/S-Z2014162 '},
        {'id': 30, 'nombre': 'UYP-CONAPROLE CIM-IC 01   Máquina-Choice Filler by Tetra Pak®         N/S-395-008-2019 '},
        {'id': 31, 'nombre': 'UYP-CONAPROLE CIM-MP 01   Máquina-Tetra Pak Aseptic Tank VD           N/S-T5844611780 '},
        {'id': 32, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-TETRA ALEX 30                       N/S-T5845523408 '},
        {'id': 33, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-TETRA ALEX 30                       N/S-5845520129 '},
        {'id': 34, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-Tetra Centri MRPX318TGV             N/S-2969461 '},
        {'id': 35, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-Tetra Centri MRPX318TGV             N/S-2969463 '},
        {'id': 36, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-Tetra Centri MRPX318TGV             N/S-2969464 '},
        {'id': 37, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-Tetra Plex MS10 SBL                 N/S-30104-38510 '},
        {'id': 38, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-Tetra Plex MS10 SBL                 N/S-30104-38511 '},
        {'id': 39, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-Tetra Plex MS6 SBL                  N/S-30104-38509 '},      
        {'id': 40, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-Tetra Plex Plex C8                  N/S-30104-38508 '},
        {'id': 41, 'nombre': 'UYP-CONAPROLE CIM-TREAT   Máquina-Tetra Plex, C8                      N/S-30107-34617 '},
        {'id': 42, 'nombre': 'UYP-CONAPROLE CIM-UHT01   Máquina-Tetra Pak Aseptic Tank LV           N/S-T5845520044 '},
        {'id': 43, 'nombre': 'UYP-CEREALIN  SJO-TREAT   Máquina-TETRA ALEX 20                       N/S-5845511263'},
        {'id': 44, 'nombre': 'UYP-CEREALIN  SJO-TREAT   Máquina-Tetra Therm Aseptic Flex 1          N/S-T5844410001'},
        {'id': 45, 'nombre': 'UYP-BARRACA   MVD-PF 01   Máquina-Tetra Alex 25                       N/S-T5856826141'},
        {'id': 46, 'nombre': 'UYP-BARRACA   MVD-PF 01   Máquina-Tetra Pak Homogenizer               N/S-5856944267'},
        {'id': 47, 'nombre': 'UYP-CONAPROLE P01-TREAT   Máquina-Tetra Plex Clip 6 RM                 N/S-30103-22150'},
        {'id': 48, 'nombre': 'UYP-CONAPROLE P01-TPPM    Máquina-Production Integrator                N/S-5861001701'},
        {'id': 49, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Tetra Pak® Separator A2              N/S-C200001'},
        {'id': 50, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Separator A2                         N/S-4269765'},
        {'id': 51, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Tetra Pak® Separator A2              N/S-C200002'},
        {'id': 52, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Separator A2                         N/S-4269764'},
        {'id': 53, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Tetra Pak® Separator H10             N/S-C200003'},
        {'id': 54, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Separator H10                        N/S-4269763'},
        {'id': 55, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Tetra Pak® Separator H714            N/S-C200004'},
        {'id': 56, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Separator H714                       N/S-4260456'},
        {'id': 57, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Tetra Pak Homogenizer                N/S-5870520020'},
        {'id': 58, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-HWU                                  N/S-D6295377/04'},
        {'id': 59, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Drum Filler                          N/S-PR2438'},
        {'id': 60, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-PHE cip                              N/S-30105-09114'},
        {'id': 61, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Skid Crema                           N/S-D6295377/01'},
        {'id': 62, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-PHE Crema                            N/S-30125-10251'},
        {'id': 63, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Skid Polisher + Desaireador          N/S-D6295377/02'},
        {'id': 64, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Tetra Plex C6-SM - AMF               N/S-30125-10250'},
        {'id': 65, 'nombre': 'UYP-CONAPROLE P07-AMF     Máquina-Tanque Pulmon                        N/S-D6295377/03'},
        {'id': 66, 'nombre': 'UYP-CONAPROLE P07-CH 01   Máquina-Tetra Pak PHE, M10                   N/S-30125-12013'},
        {'id': 67, 'nombre': 'UYP-CONAPROLE P07-TREAT   Máquina-Tetra Plex Clip 10 RH                N/S-30103-17680'},
        {'id': 68, 'nombre': 'UYP-CONAPROLE P07-TREAT   Máquina-Tetra Plex Clip 6 RM                 N/S-30103-17688'},
        {'id': 70, 'nombre': 'UYP-CONAPROLE P07-TREAT   Máquina-Tetra Centri HMRPX718HGV74C          N/S-4110484'},
        {'id': 71, 'nombre': 'UYP-CONAPROLE P07-TREAT   Máquina-Tetra Centri HMRPX718HGV74C          N/S-4110489'},
        {'id': 72, 'nombre': 'UYP-CONAPROLE P07-TREAT   Máquina-Tetra Plex C6                        N/S-30104-21165'},
        {'id': 73, 'nombre': 'UYP-CONAPROLE P07-TREAT   Máquina-Tetra Plex C8                        N/S-30104-21168'},
        {'id': 74, 'nombre': 'UYP-CONAPROLE P07-TREAT   Máquina-Tetra Plex M10 M BASE                N/S-30103-17682'},
        {'id': 75, 'nombre': 'UYP-CONAPROLE P07-TREAT   Máquina-Tetra Plex Clip 10 RM                N/S-30103-17681'},
        {'id': 76, 'nombre': 'UYP-CONAPROLE P08-DR 01   Máquina-Pasteurizador 4ta línea estandarización       N/S-D6333928/01'},
        {'id': 77, 'nombre': 'UYP-CONAPROLE P08-DR 01   Máquina-Tetra Pak PHE, C15 Sanitary                   N/S-30125-35464'},
        {'id': 78, 'nombre': 'UYP-CONAPROLE P08-DR 01   Máquina-Tetra Pak PHE, M6                             N/S-30125-11940'},
        {'id': 79, 'nombre': 'UYP-CONAPROLE P08-DR 01   Máquina-Tetra Pak Standardization unit                N/S-T5845470166'},
        {'id': 80, 'nombre': 'UYP-CONAPROLE P08-DR 01   Máquina-Tetra Pak® Separator H80                      N/S-C220009'},
        {'id': 81, 'nombre': 'UYP-CONAPROLE P08-DR 01   Máquina-Separator H80                                  N/S-AAF0000430'},
        {'id': 82, 'nombre': 'UYP-CONAPROLE P08-MX 01   Máquina-Tetra Pak High Shear Mixer                     N/S-T5845706418'},
        {'id': 83, 'nombre': 'UYP-CONAPROLE P08-MX 01   Máquina-Tetra Pak High Shear Mixer                     N/S-T5845706420'},
        {'id': 84, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Powder Handling                      N/S-5860700023'},
        {'id': 85, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Big Bag Tipping unit VB005          N/S-5860010176'},
        {'id': 86, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Big Bag Tipping unit VB005          N/S-5860010177'},
        {'id': 87, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040608'},
        {'id': 88, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040609'},
        {'id': 89, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Lump Breaker TD041                  N/S-5860190027'},
        {'id': 90, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Bag Tipping unit M VS401           N/S-5860020148'},
        {'id': 91, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Dust Filter C FL002                 N/S-5860170368'},
        {'id': 92, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Rotary Valve FB                     N/S-5860070145'},
        {'id': 93, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Rotary Valve FB                     N/S-5860070146'},
        {'id': 94, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®  Inlet Air Managment PLP          N/S-5860240325'},
        {'id': 95, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®  Outlet Air Managment PLP         N/S-5860240326'},
        {'id': 96, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280654'},
        {'id': 97, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040610'},
        {'id': 98, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040611'},
        {'id': 99, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280655'},
        {'id': 100, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280656'},
        {'id': 101, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280657'},
        {'id': 102, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280658'},
        {'id': 103, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040612'},
        {'id': 104, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040613'},
        {'id': 105, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040614'},
        {'id': 106, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040615'},
        {'id': 107, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                   N/S-5860040616'},
        {'id': 108, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®  Inlet Air Managment PLP          N/S-5860240327'},
        {'id': 109, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®  Outlet Air Managment PLP         N/S-5860240328'},
        {'id': 110, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Rotary Valve FB                    N/S-5860070147'},
        {'id': 111, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Rotary Valve FB                    N/S-5860070148'},
        {'id': 112, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Dust Filter C FL002                N/S-5860170369'},
        {'id': 113, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Dust Filter C FL002                N/S-5860170370'},
        {'id': 114, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280659'},
        {'id': 115, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280660'},
        {'id': 116, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280661'},
        {'id': 117, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280662'},
        {'id': 118, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Big Bag Filling unit S DB006       N/S-5860140027'},
        {'id': 119, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Bag Tipping unit CS VS600          N/S-5860310077'},
        {'id': 120, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Bag Tipping unit CS VS600          N/S-5860310078'},
        {'id': 121, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Bag Tipping unit CS VS600          N/S-5860310079'},
        {'id': 122, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Bag Tipping unit CS VS600          N/S-5860310080'},
        {'id': 123, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280663'},
        {'id': 124, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280664'},
        {'id': 125, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280665'},
        {'id': 126, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Surge Hopper VT                    N/S-5860280666'},
        {'id': 127, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Lump Breaker TD041                 N/S-5860190030'},
        {'id': 128, 'nombre': 'UYP-CONAPROLE P08-PP 01   Máquina-Tetra Pak®Screw Conveyor ET                  N/S-5860040689'},
        {'id': 129, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Centri MRPX214TGV44C                   N/S-2917438'},
        {'id': 130, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Centri MRPX214TGV44                    N/S-2985490'},
        {'id': 131, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Centri MRPX214-74T                     N/S-2893664'},
        {'id': 132, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Centri MRPX214-TGV74                   N/S-2898316'},
        {'id': 133, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Alfast 220                             N/S-T5845420205'},
        {'id': 134, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Plex Clip 10 RM                        N/S-30103-22146'},
        {'id': 135, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Therm Lacta                            N/S-D00623-01/1'},
        {'id': 136, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Plex M10 B                             N/S-30100-22169'},
        {'id': 137, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Plex M6 M                              N/S-30103-22147'},
        {'id': 139, 'nombre': 'UYP-CONAPROLE P08-TREAT   Máquina-Tetra Plex M6 MFMC                           N/S-30100-41336'},
        {'id': 140, 'nombre': 'UYP-CONAPROLE P08-UHT01   Máquina-Tetra Pak Aseptic Tank LV                    N/S-T5845520045'},
        {'id': 141, 'nombre': 'UYP-CONAPROLE P08-UHT01   Máquina-Tank body Tetra Pak Aseptic Tank VD/LV       N/S-211182'},
        {'id': 142, 'nombre': 'UYP-CONAPROLE P08-UHT01   Máquina-Tetra Pak Aseptic Dosing unit F              N/S-T5844550349'},
        {'id': 143, 'nombre': 'UYP-CONAPROLE P08-TPPM    Máquina-Production Control                           N/S-5861001994'},
        {'id': 144, 'nombre': 'UYP-CONAPROLE P08-MP 01   Máquina-Tetra Therm Lacta                            N/S-T5845140314'},
        {'id': 145, 'nombre': 'UYP-CONAPROLE P08-MP 01   Máquina-Tetra Plex, C10                              N/S-30109-16965'},
        {'id': 146, 'nombre': 'UYP-CONAPROLE P08-MP 01   Máquina-Tetra Pak Homogenizer                        N/S-5871117115'},
        {'id': 147, 'nombre': 'UYP-CONAPROLE P09-CH 01   Máquina-Tetra Tebel Ost IV                           N/S-620910.01'},
        {'id': 148, 'nombre': 'UYP-CONAPROLE P09-CH 01   Máquina-Tetra Tebel Ost IV                           N/S-620910.02'},
        {'id': 149, 'nombre': 'UYP-CONAPROLE P09-CH 01   Máquina-Tetra Tebel Ost IV                           N/S-620910.03'},
        {'id': 150, 'nombre': 'UYP-CONAPROLE P09-CH 01   Máquina-Tetra Tebel Ost IV                           N/S-620910.04'},
        {'id': 151, 'nombre': 'UYP-CONAPROLE P09-CH 01   Máquina-Tetra Tebel Alfomatic 2B                     N/S-620922.01'},
        {'id': 152, 'nombre': 'UYP-CONAPROLE P09-CH 01   Máquina-Tetra Tebel OST SH 5.1 12.500 L              N/S-5847110077'},
        {'id': 153, 'nombre': 'UYP-CONAPROLE P09-CH 01   Máquina-Tetra Damrow Double OO CH 8.0 14.000L        N/S-5847190007'},
        {'id': 154, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Alfast 220                             N/S-T5845420207'},
        {'id': 155, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Alfast 220                             N/S-T5845420206'},
        {'id': 156, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Plex Clip 10 RM                        N/S-30103-22144'},
        {'id': 157, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Centri HMRPX414HGV74C                  N/S-417117'},
        {'id': 159, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Centri HMRPX614HGV74C                  N/S-4042654'},
        {'id': 160, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Therm Lacta                            N/S-D00623-01/3'},
        {'id': 161, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Plex M6 M                              N/S-30103-22149'},
        {'id': 162, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Plex C8-KSR                            N/S-30106-34564'},
        {'id': 163, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Plex, C8                               N/S-30109-15185'},
        {'id': 164, 'nombre': 'UYP-CONAPROLE P09-TREAT   Máquina-Tetra Plex, M10                              N/S-30109-15186'},
        {'id': 165, 'nombre': 'UYP-CONAPROLE P11-TREAT   Máquina-Tetra Centri HMRPX514HGV74C                  N/S-4033885'},
        {'id': 166, 'nombre': 'UYP-CONAPROLE P11-TREAT   Máquina-Tetra Centri HMRPX714HGV74C                  N/S-4047653'},
        {'id': 167, 'nombre': 'UYP-CONAPROLE P16-TREAT   Máquina-Tetra Plex Clip 10 RM                        N/S-30103-22145'},
        {'id': 168, 'nombre': 'UYP-CONAPROLE P16-TREAT   Máquina-Tetra Therm Lacta                            N/S-D00623-01/2'},
        {'id': 169, 'nombre': 'UYP-CONAPROLE P16-TREAT   Máquina-Tetra Plex M6 M                              N/S-30103-22148'},
        {'id': 170, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Contisolv                         	  N/S-15735-01'},
        {'id': 171, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco CIP unit                          	  N/S-15735-02'},
        {'id': 172, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco CIP unit                          	  N/S-15735-03'},
        {'id': 173, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Thermizer                         	  N/S-15735-04'},
        {'id': 174, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco CIP unit                          	  N/S-15735-05'},
        {'id': 175, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Ion Exchange System                   N/S-15735-06'},
        {'id': 176, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Horizontal Plate Filtration           N/S-16676-01'},
        {'id': 177, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco RJ Mixer                              N/S-16958-02'},
        {'id': 178, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco RJ Mixer                              N/S-16958-03'},
        {'id': 179, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco RJ Mixer                              N/S-16958-04'},
        {'id': 180, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco RJ Mixer                              N/S-16958-05'},
        {'id': 181, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Ingredients Dissolver                 N/S-15229-01'},
        {'id': 182, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Filtration                            N/S-15735-07'},
        {'id': 183, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Filtration                            N/S-15735-08'},
        {'id': 184, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Ingredients Dissolver                 N/S-16676-'},
        {'id': 185, 'nombre': 'UYP-MONRESA   MVD-MITEC   Máquina-Miteco Ingredients Dissolver                 N/S-T5880160029'},
        {'id': 186, 'nombre': 'UYP-MONRESA   MVD-MX01    Máquina-Tetra Pak Mixer RJCI                         N/S-T5845680002'},
        {'id': 187, 'nombre': 'UYC-JUGOSURUG FRY-LIN01   Máquina- TBA/21                                      N/S-15011/00099'},
        {'id': 188, 'nombre': 'UYC-JUGOSURUG FRY-LIN02   Máquina- A3/Flex                                     N/S-21211/00147'},
        {'id': 189, 'nombre': 'UYC-JUGOSURUG FRY-LIN03   Máquina- A3/Flex                                     N/S-21211/00389'},
        {'id': 190, 'nombre': 'UYP-JUGOSURUG FRY-TREAT   Máquina-Tetra Plex Clip 6 RM WS                      N/S-30101-22606'},
        {'id': 191, 'nombre': 'UYP-JUGOSURUG FRY-TREAT   Máquina-Tetra Plex CLIP 6 WS                         N/S-30101-25950'},
        {'id': 192, 'nombre': 'UYP-JUGOSURUG FRY-TREAT   Máquina-Tetra Therm Nacional                         N/S-95085'},
        {'id': 193, 'nombre': 'UYP-JUGOSURUG FRY-TREAT   Máquina-Tetra Therm Nacional                         N/S-B97439-01'},
        {'id': 194, 'nombre': 'UYP-JUGOSURUG FRY-Z_DOC   Máquina-Plant Documentation                          N/S-D5865000135'},
        {'id': 195, 'nombre': 'UYC-CONAPROLE P08-LIN05   Máquina-E3/Speed                                     N/S-21225/00033'},
        {'id': 196, 'nombre': 'UYP-ECOLAT    NVH-TREAT TETRA ALEX 30 A   Máquina-T4625220                     N/S-T4625220'},
        {'id': 197, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Alsafe LA   Máquina-T2440170                     N/S-T2440170'},
        {'id': 198, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Therm Aseptic Flex   Máquina-T2100197            N/S-T2100197'},
        {'id': 199, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Alfast Plus   Máquina-T5845440334                N/S-T5845440334'},
        {'id': 200, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Plex MS6-SR in C6092071/04   Máquina-30106-35480 N/S-30106-35480'},
        {'id': 201, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Plex C6-SR in C6092071/01   Máquina-30106-35479  N/S-30106-35479'},
        {'id': 202, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Alcip Nacional   Máquina-C6092071/02             N/S-C6092071/02'},
        {'id': 203, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Therm Nacional   Máquina-C6092071/04             N/S-C6092071/04'},
        {'id': 204, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Therm Nacional   Máquina-C6092071/01             N/S-C6092071/01'},
        {'id': 205, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Therm Nacional   Máquina-C6092071/03             N/S-C6092071/03'},
        {'id': 206, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Plex MS6-SR in C6092071/03   Máquina-30106-35573 N/S-30106-35573'},
        {'id': 207, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Plex C6-SR   Máquina-30106-35478                 N/S-30106-35478'},
        {'id': 208, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Plex C6-SR   Máquina-30106-35477                 N/S-30106-35477'},
        {'id': 209, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Alcross US 100   Máquina-T5845810090             N/S-T5845810090'},
        {'id': 210, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Plex MS10-SR   Máquina-30107-34354               N/S-30107-34354'},
        {'id': 211, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Alfast   Máquina-T5845440556                     N/S-T5845440556'},
        {'id': 212, 'nombre': 'UYP-ECOLAT    NVH-TREAT Tetra Alfast   Máquina-T5845440557                     N/S-T5845440557'},
        {'id': 213, 'nombre': 'UYP - N/A'},
        {'id': 214, 'nombre': 'UYC - N/A'}
    ]

    if request.method == 'POST':
        try:
            fecha = request.form['fecha']
            entrada = f"{request.form['entrada_hora']}:{request.form['entrada_minuto']}"
            salida = f"{request.form['salida_hora']}:{request.form['salida_minuto']}"


            almuerzo_horas = int(request.form.get('almuerzo_horas', 0))
            almuerzo = almuerzo_horas

            viaje_ida = int(request.form.get('viaje_ida', 0))
            viaje_vuelta = int(request.form.get('viaje_vuelta', 0))
            km_ida = int(request.form.get('km_ida', 0))
            km_vuelta = int(request.form.get('km_vuelta', 0))

            if viaje_ida < 0 or viaje_vuelta < 0 or km_ida < 0 or km_vuelta < 0:
                flash("Los valores numéricos no pueden ser negativos", "danger")
                return redirect(url_for('editar_registro', id=id))

            tarea = request.form.get('tarea', '')
            cliente = request.form.get('cliente', '')
            contrato = request.form.get('contrato')
            service_order = request.form.get('service_order', '')
            centro_costo = request.form.get('centro_costo', '')
            tipo_servicio = request.form.get('tipo_servicio', '')
            linea = request.form.get('linea', '')
            comentarios = request.form.get('comentarios', '')

            t_entrada = datetime.strptime(entrada, "%H:%M")
            t_salida = datetime.strptime(salida, "%H:%M")

            horas_trabajadas = (t_salida - t_entrada - timedelta(hours=almuerzo)).total_seconds() / 3600
        except ValueError:
            flash("Por favor, ingresá valores válidos y formatos correctos", "danger")
            return redirect(url_for('editar_registro', id=id))

        # Guardar en el registro
        registro.fecha = fecha
        registro.entrada = entrada
        registro.salida = salida
        registro.almuerzo = round(almuerzo, 2)
        registro.viaje_ida = viaje_ida
        registro.viaje_vuelta = viaje_vuelta
        registro.km_ida = km_ida
        registro.km_vuelta = km_vuelta
        registro.horas = round(horas_trabajadas, 2)
        registro.tarea = tarea
        registro.cliente = cliente
        registro.contrato = contrato  # ✅ correcto  # guarda el valor seleccionado, como '73450003'
        registro.service_order = service_order
        registro.centro_costo = centro_costo
        registro.tipo_servicio = tipo_servicio
        registro.linea = linea
        registro.comentarios = comentarios

        db.session.commit()
        flash('Registro actualizado exitosamente', 'success')

        contexto = request.args.get('contexto', 'admin')  # valor por defecto

        if session.get('role') in ['admin', 'superadmin']:
            if contexto == 'dashboard':
                return redirect(url_for('dashboard'))
            else:
                return redirect(url_for('admin', filtro_usuario=filtro_usuario, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta))
        else:
            return redirect(url_for('dashboard'))


        
    cliente_prefijo = {
        '1561046 - Barraca Deambrosi SA (Carton)'   : 'UYC-BARRACA',
        '1561046 - Barraca Deambrosi SA (Proceso)'  : 'UYP-BARRACA',
        '1561458 - Belficor Sociedad Anónima (Proceso)'       : 'UYP-BELFICOR',
        '1561459 - Belficor - Paraje La Boyada'     : 'UYP-BELFICOR',
        '1561233 - Companía Salus SA'               : 'UYP-FTE SALUS',
        '1561251 - Cooperativa Agraria (CALCAR)'    : 'UYC-COAGRARIA',
        '1560621 - Ecolat Uruguay S.A. (Proceso)'   : 'UYP-ECOLAT',
        '1561335 - Gibur S.A. (Carton)'             : 'UYC-GIBUR',
        '1561540 - Nolir S.A.'                      : 'UYC-NOLIR',
        '1560020 - Recalco S.A. (ex Suadil)'        : 'UYC-RECALCO',
        '1561245 - Conaprole Planta I (Proceso)'    : 'UYP-CONAPROLE P01',
        '1561250 - Conaprole Planta CIM (Carton)'   : 'UYC-CONAPROLE CIM',
        '1561250 - Conaprole Planta CIM (Proceso)'  : 'UYP-CONAPROLE CIM',
        '1561010 - Conaprole Planta VII (Proceso)'  : 'UYP-CONAPROLE P07',
        '1560606 - Conaprole Planta VIII (Carton)'  : 'UYC-CONAPROLE P08',
        '1560606 - Conaprole Planta VIII (Proceso)' : 'UYP-CONAPROLE P08',
        '1560018 - Cerealin San Jose (Carton)'      : 'UYC-CEREALIN',
        '1560018 - Cerealin San Jose (Proceso)'     : 'UYP-CEREALIN',
        '1569004 - Cerealin S.A.'          : 'UYP-CEREALIN',
        '1561132 - Montevideo Refrescos S.A. (Proceso)': 'UYP-MONRESA',
        '1561062 - Jugos del Uruguay S.A. (Carton)'   : 'UYC-JUGOSURUG',  # definir si hay prefijo
        '1561062 - Jugos del Uruguay S.A. (Proceso)'  : 'UYP-JUGOSURUG',  # definir si hay prefijo
        '1570011 - Tetrapak Argentina'           : 'N/A',
        'Otros Clientes del Mercado'      : 'UYP - N/A',
        'Otros Clientes del Mercado'      : 'UYC - N/A'
    }
      # ─── Construcción automática de cliente_cc_lineas ───
    
    cliente_cc_lineas = {}
    for cli in clientes:
        centros = [cc['nombre'] for cc in centros_costo if cli in cc['nombre']]
        pref = cliente_prefijo.get(cli, '')
        if pref:
            lineas_f = [ln['nombre'] for ln in lineas if ln['nombre'].startswith(pref)]
        else:
            lineas_f = []
        cliente_cc_lineas[cli] = {
            'centros_costo': centros,
            'lineas':        lineas_f
        }    
        
    role = session.get('role')  # 👈 esto es lo que falta
    contexto = request.args.get('contexto', 'admin')  # 👈 esta línea
    # GET: mostrar formulario con datos y listas para selects
    contrato_labels = {item['value']: item['label'] for item in contratos}

    return render_template('editar_registro.html',
                           registro=registro,
                           lista_clientes=clientes,
                           contratos=contratos,
                           contrato_labels=contrato_labels,
                           #contratos=[{'nombre': c} for c in contratos], // Original
                           service_orders=service_orders,
                           centros_costo=centros_costo,
                           tipos_servicio=tipos_servicio,
                           lineas=lineas,
                           cliente_cc_lineas = cliente_cc_lineas,
                           role=role,
                           contexto=contexto  # 👈 esto habilita los campos en el template
    )


@app.route('/borrar_registro/<int:id>', methods=['POST'])
def borrar_registro(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    registro = Registro.query.get_or_404(id)
    db.session.delete(registro)
    db.session.commit()
    return redirect(url_for('admin') if session['role'] == 'superadmin' else url_for('dashboard'))

@app.route('/crear_admin', methods=['GET', 'POST'])
def crear_admin():
    if 'user_id' not in session or session['role'] != 'superadmin':
        return redirect(url_for('login'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password')
        confirmar = request.form.get('confirmar_password')

        if not username or not password or not confirmar:
            flash('Todos los campos son obligatorios.', category='warning')
            return render_template('crear_admin.html')

        if password != confirmar:
            flash('Las contraseñas no coinciden.', category='danger')
            return render_template('crear_admin.html')

        if User.query.filter_by(username=username).first():
            flash('Ese nombre de usuario ya existe.', category='danger')
        else:
            nuevo_admin = User(username=username, password=password, role='admin')
            db.session.add(nuevo_admin)
            db.session.commit()
            flash('Administrador creado correctamente', category='success')

    return render_template('crear_admin.html')


@app.route('/administrator', methods=['GET'])
def admin():
    if 'user_id' not in session or session['role'] not in ['admin', 'superadmin']:
        return redirect(url_for('login'))

    # Obtener usuarios para el filtro
    usuarios = User.query.with_entities(User.id, User.username).all()

    # Obtener filtros desde GET (formulario usa método GET)
    filtro_usuario = request.args.get('filtro_usuario')
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')

    # Query base
    query = db.session.query(Registro, User).join(User)

    # Filtro por usuario
    # Filtro por usuario (convertimos a entero con manejo de errores)
    filtro_usuario = request.args.get('filtro_usuario')
    if filtro_usuario:
        try:
            filtro_usuario = int(filtro_usuario)
            query = query.filter(User.id == filtro_usuario)
        except ValueError:
            filtro_usuario = None  # Si el valor no es un número válido


    # Filtro por fechas
    if fecha_desde and fecha_hasta:
        query = query.filter(Registro.fecha.between(fecha_desde, fecha_hasta))

    registros = query.order_by(Registro.fecha.desc()).all()
    # Definir contratos
    contratos = [
        {'value': '73450003', 'label': 'Contrato'},
        {'value': '79010000', 'label': 'Administrativo - 79010000'},
        {'value': '79200030', 'label': 'Issue Resolution - Service - 79200030'},
        {'value': '79200020', 'label': 'Issue Resolution - Capital CARTON - 79200020'},
        {'value': '79200050', 'label': 'Issue Resolution - Capital PROCESSING - 79200050'},
        {'value': '580000', 'label': 'Capital Equipment carton - 580000'},
        {'value': '70350000', 'label': 'Capital Equipment processing - 70350000'},
        {'value': '73450000', 'label': 'Fuera de contrato - 73450000'},
        {'value': '560000', 'label': 'Upgrade projects - 560000'},
        {'value': '70158000', 'label': 'Mandatory kit - 70158000'},
        {'value': '480000', 'label': 'Rebilling - 480000'},
        {'value': '73453000', 'label': 'Training Facturable - 73453000'},
        {'value': '79010000', 'label': 'Training Interno - 79010000'},
        {'value': '70400010', 'label': 'MDR - 70400010'},
        {'value': '79010000', 'label': 'No disponible - 79010000'},
        {'value': '0', 'label': 'N/A'}
    ]

    contrato_labels = {item['value']: item['label'] for item in contratos}

    return render_template(
        'admin.html',
        registros=registros,
        usuarios=usuarios,
        filtro_usuario=filtro_usuario,
        contratos=contratos,
        contrato_labels=contrato_labels,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        username=session['username'],
        role=session['role']
    )



@app.route('/cambiar_password', methods=['GET', 'POST'])
def cambiar_password():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        nueva = request.form['nueva']
        confirmar = request.form['confirmar']  # Se agrega para la comparación de contraseñas

        if nueva != confirmar:
            flash('Las contraseñas no coinciden.', category='danger')
            return render_template('cambiar_password.html')

        # Si las contraseñas coinciden, actualizarla en la base de datos
        user = User.query.get(session['user_id'])
        user.password = nueva
        db.session.commit()
        flash('Contraseña actualizada', category='success')

    return render_template('cambiar_password.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/crear_usuario', methods=['GET', 'POST'])
def crear_usuario():
    if 'user_id' not in session or session['role'] not in ['admin', 'superadmin']:
        return redirect(url_for('login'))

    if request.method == 'POST':
        username = request.form['username'].strip().lower()
        password = request.form['password']
        confirmar = request.form['confirmar_password']

        if password != confirmar:
            flash('Las contraseñas no coinciden.', category='danger')
            return render_template('crear_usuario.html')

        if User.query.filter_by(username=username).first():
            flash('Ese nombre de usuario ya existe.', category='danger')
        else:
            nuevo_usuario = User(username=username, password=password, role='usuario')
            db.session.add(nuevo_usuario)
            db.session.commit()
            flash('Usuario creado exitosamente.', category='success')


    return render_template('crear_usuario.html')

@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        username = request.form['username'].strip().lower()
        password = request.form['password']
        confirmar = request.form['confirmar_password']

        if password != confirmar:
            flash('Las contraseñas no coinciden.', category='danger')
            return render_template('registro.html')

        if User.query.filter_by(username=username).first():
            flash('Ese nombre de usuario ya existe.' , category='danger')
        else:
            nuevo_usuario = User(username=username, password=password, role='usuario')
            db.session.add(nuevo_usuario)
            db.session.commit()
            flash('Usuario creado exitosamente. Ahora podés iniciar sesión.', category='success')
            return redirect(url_for('login'))

    return render_template('registro.html')

@app.route('/lista_usuarios')
def lista_usuarios():
    if 'user_id' not in session or session.get('role') != 'superadmin':
        return redirect(url_for('login'))

    usuarios = User.query.all()
    return render_template('usuarios.html', usuarios=usuarios)


@app.route('/editar_usuario/<int:id>', methods=['GET', 'POST'])
def editar_usuario(id):
    if 'user_id' not in session or session['role'] != 'superadmin':
        return redirect(url_for('login'))  # Si no es superadmin, redirigir al login

    user = User.query.get_or_404(id)  # Busca el usuario por ID, si no lo encuentra, lanza error 404

    if request.method == 'POST':  # Si se recibe una solicitud POST (cuando el formulario es enviado)
        user.username = request.form['username'].strip().lower()
        #user.email = request.form['email']
        user.role = request.form['role']
        user.password = request.form['password']  # <- Agregado acá
        db.session.commit()  # Realiza el commit en la base de datos para guardar los cambios
        flash('Usuario actualizado correctamente', 'success')  # Mensaje de éxito
        return redirect(url_for('lista_usuarios'))  # Redirige a la lista de usuarios después de la edición

    return render_template('editar_usuarios.html', user=user)  # Si es GET, muestra el formulario con los datos actuales

@app.route('/eliminar_usuario/<int:id>', methods=['POST'])
def eliminar_usuario(id):
    if 'user_id' not in session or session['role'] != 'superadmin':
        return redirect(url_for('login'))

    user = User.query.get_or_404(id)
    db.session.delete(user)
    db.session.commit()
    flash('Usuario eliminado correctamente', 'danger')
    return redirect(url_for('lista_usuarios'))  # Cambio aquí


if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
